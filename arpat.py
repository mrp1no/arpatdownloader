import urllib.request
import openpyxl
from pyproj import CRS, Transformer
from datetime import date
now = date.today()
filename = now.strftime("%Y-%m-%d")
print(str(filename) + " | ARPAT Downloader v1.0\n\nConnessione a sira.arpat.toscana.it...")
opener = urllib.request.urlopen('http://sira.arpat.toscana.it/sira/misure_rf/portale.php')
print('Fatto\n\nCaricamento in corso...')
content = opener.read()
print('Fatto\n\nEstrazione dati...')
def textdel(orig, frase):
    indice = orig.find(frase)
    if indice != -1:
        return orig[indice:]
    else:
        return orig
content2 = textdel(str(content), "Provincia|Comune|Indirizzo|LOCALITA|COORDINATA_EST|COORDINATA_NORD|ALTITUDINE|Tipologia|Gestore|RAGIONE_SOCIALE_Gestore|Nome|TIPOLOGIA_IMPIANTI|Tecnologia|Riferimento")
def textdel2(orig2, frase2):
    index = orig2.find(frase2)
    if index != -1:
        orig2 = orig2[:index]
    return orig2
content3 = textdel2(content2, '","Provincia|Comune|Indirizzo|COORDINATA_EST|COORDINATA_NORD|Tipologia|Data|Valore misurato (V/m)')
content4 = content3.replace("!n!", "\n")
def crea_excel(testo, nome_file):
    righe = testo.split("\n")
    foglio_excel = openpyxl.Workbook()
    foglio_attivo = foglio_excel.active
    
    for riga in righe:
        celle = riga.split("|")
        foglio_attivo.append(celle)
    
    foglio_excel.save(nome_file)

crea_excel(str(content4), filename + ".xlsx")

gauss_boaga_crs = CRS.from_epsg(3003)
wgs84_crs = CRS.from_epsg(4326)

# Creazione di un trasformatore per la conversione delle coordinate
transformer = Transformer.from_crs(gauss_boaga_crs, wgs84_crs)

# Apertura del file di lavoro
workbook = openpyxl.load_workbook(filename + '.xlsx')

# Selezione del foglio di lavoro
sheet = workbook.active

# Iterazione sulle righe del foglio
for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    est = row[4]  # Column 5 (zero-based index)
    nord = row[5]  # Column 6 (zero-based index)

    # Conversione coordinate da Gauss Boaga a WGS84
    lon, lat = transformer.transform(est, nord)

    # Aggiornamento delle colonne con le coordinate WGS84
    sheet.cell(row=i, column=6, value=lat)
    sheet.cell(row=i, column=5, value=lon)
    
# Itera sulle righe del foglio
for row in sheet.iter_rows(min_row=2):  # Inizia dalla riga 2 perché nella prima c'è il titolo
    # Sostituzione accenti e virgole delle coordinate
    row[1].value = str(row[1].value).replace("\\xc2\\xb0", "°")  # Colonna 2
    row[1].value = str(row[1].value).replace("\\xc3\\xb9", "ù")  # Colonna 2
    row[1].value = str(row[1].value).replace("\\xc3\\xb2", "ò")  # Colonna 2
    row[1].value = str(row[1].value).replace("\\xc3\\xa0", "à")  # Colonna 2
    row[1].value = str(row[1].value).replace("\\xc3\\x80", "À")  # Colonna 2
    row[1].value = str(row[1].value).replace("\\xc3\\xa8", "è")  # Colonna 2
    row[1].value = str(row[1].value).replace("\\", "")  # Colonna 2
    row[2].value = str(row[2].value).replace("\\xc2\\xb0", "°")  # Colonna 3
    row[2].value = str(row[2].value).replace("\\xc2\\xa0", "")  # Colonna 3
    row[2].value = str(row[2].value).replace("\\xc3\\xa8", "è")  # Colonna 3
    row[2].value = str(row[2].value).replace("\\xc3\\xa0", "à")  # Colonna 3
    row[2].value = str(row[2].value).replace("\\xc3\\x88", "È")  # Colonna 3
    row[2].value = str(row[2].value).replace("\\xc3\\xb2", "ò")  # Colonna 3
    row[2].value = str(row[2].value).replace("\\xc3\\xa9", "é")  # Colonna 3
    row[2].value = str(row[2].value).replace("\\xc3\\x80", "À")  # Colonna 3
    row[2].value = str(row[2].value).replace("\\xc3\\xb9", "ù")  # Colonna 3
    row[2].value = str(row[2].value).replace("\\", "")  # Colonna 3
    row[3].value = str(row[3].value).replace("\\xc3\\x80", "À")  # Colonna 4
    row[3].value = str(row[3].value).replace("\\xc3\\xa0", "à")  # Colonna 4
    row[3].value = str(row[3].value).replace("\\xc3\\xa9", "é")  # Colonna 4
    row[3].value = str(row[3].value).replace("\\xc2\\xb0", "°")  # Colonna 4
    row[3].value = str(row[3].value).replace("\\", "")  # Colonna 4
    row[4].value = str(row[4].value).replace(',', '.')  # Colonna 5
    row[5].value = str(row[5].value).replace(',', '.')  # Colonna 6
    row[8].value = str(row[8].value).replace("\\xc3\\xb9", "ù")  # Colonna 9
    row[8].value = str(row[8].value).replace("\\xc3\\xa0", "à")  # Colonna 9
    row[9].value = str(row[9].value).replace("\\xc3\\xb9", "ù")  # Colonna 10
    row[9].value = str(row[9].value).replace("\\xc3\\xa0", "à")  # Colonna 10
    row[9].value = str(row[9].value).replace("\\", "")  # Colonna 10
    row[10].value = str(row[10].value).replace("\\xc2\\xb0", "°")  # Colonna 11
    row[10].value = str(row[10].value).replace("\\xc3\\xb7", "÷")  # Colonna 11
    row[10].value = str(row[10].value).replace("\\xc2\\xb4", "'")  # Colonna 11
    row[10].value = str(row[10].value).replace("\\xc3\\xa0", "à")  # Colonna 11
    row[10].value = str(row[10].value).replace("\\xc3\\x88", "È")  # Colonna 11
    row[10].value = str(row[10].value).replace("\\xc3\\xb2", "ò")  # Colonna 11
    row[10].value = str(row[10].value).replace("\\xc3\\xa8", "è")  # Colonna 11
    row[10].value = str(row[10].value).replace("\\xc3\\x80", "À")  # Colonna 11
    row[10].value = str(row[10].value).replace("\\xc3\\xb9", "ù")  # Colonna 11
    row[10].value = str(row[10].value).replace("\\", "")  # Colonna 11

# Export finale
workbook.save(filename + '.xlsx')
print('Fatto\n\nL\'export è stato salvato come \'' + filename +  '.xlsx\' nella directory dello script.\n')